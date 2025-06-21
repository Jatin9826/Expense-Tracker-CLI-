
import json
import os
import datetime
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox, simpledialog

# Filenames
JSON_FILE = "expenses.json"
EXCEL_FILE = "expenses.xlsx"

# ---------------------- Data Handling ---------------------- #
def load_expenses():
    if not os.path.exists(JSON_FILE):
        return []
    with open(JSON_FILE, 'r') as f:
        return json.load(f)

def save_expenses(expenses):
    with open(JSON_FILE, 'w') as f:
        json.dump(expenses, f, indent=4)

def write_to_excel(expense):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Amount", "Category", "Date"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    ws.append([expense["amount"], expense["category"], expense["date"]])
    wb.save(EXCEL_FILE)

# ---------------------- Expense Functions ---------------------- #
def add_expense_gui():
    try:
        amount = float(entry_amount.get())
        category = entry_category.get()
        date = entry_date.get()

        if not date:
            date_val = datetime.date.today()
        else:
            date_val = datetime.datetime.strptime(date, "%Y-%m-%d").date()

        expense = {
            "amount": amount,
            "category": category,
            "date": str(date_val)
        }

        expenses.append(expense)
        save_expenses(expenses)
        write_to_excel(expense)

        messagebox.showinfo("Success", "✅ Expense added!")
        entry_amount.delete(0, tk.END)
        entry_category.delete(0, tk.END)
        entry_date.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Error", f"❌ {e}")

def view_expenses_gui():
    if not expenses:
        messagebox.showinfo("Expenses", "No expenses recorded yet.")
        return
    result = "\n".join([f"₹{exp['amount']} | {exp['category']} | {exp['date']}" for exp in expenses])
    messagebox.showinfo("📋 All Expenses", result)

def show_summary_gui():
    try:
        month = simpledialog.askinteger("Input", "Enter month (1-12):")
        year = simpledialog.askinteger("Input", "Enter year (e.g., 2025):")

        filtered = [
            exp for exp in expenses
            if datetime.datetime.strptime(exp['date'], "%Y-%m-%d").month == month
            and datetime.datetime.strptime(exp['date'], "%Y-%m-%d").year == year
        ]

        if not filtered:
            messagebox.showinfo("Summary", "No expenses found for the given month.")
            return

        totals = defaultdict(float)
        for exp in filtered:
            totals[exp['category']] += exp['amount']

        summary = "\n".join([f"{cat}: ₹{amt:.2f}" for cat, amt in totals.items()])
        messagebox.showinfo(f"📊 Summary for {month}/{year}", summary)

        # Plot pie chart
        plt.figure(figsize=(6, 6))
        plt.pie(totals.values(), labels=totals.keys(), autopct='%1.1f%%', startangle=140)
        plt.title(f"Expenses Breakdown for {month}/{year}")
        plt.axis('equal')
        plt.show()

    except Exception as e:
        messagebox.showerror("Error", f"❌ {e}")

# ---------------------- GUI Setup ---------------------- #
expenses = load_expenses()

app = tk.Tk()
app.title("💰 Expense Tracker")
app.geometry("300x320")
app.config(padx=10, pady=10)

tk.Label(app, text="Amount (₹):").pack()
entry_amount = tk.Entry(app)
entry_amount.pack()

tk.Label(app, text="Category:").pack()
entry_category = tk.Entry(app)
entry_category.pack()

tk.Label(app, text="Date (YYYY-MM-DD, optional):").pack()
entry_date = tk.Entry(app)
entry_date.pack()

tk.Button(app, text="Add Expense", command=add_expense_gui, bg="#28a745", fg="white").pack(pady=5)
tk.Button(app, text="View Expenses", command=view_expenses_gui, bg="#007bff", fg="white").pack(pady=5)
tk.Button(app, text="Monthly Summary with Chart", command=show_summary_gui, bg="#ffc107").pack(pady=5)
tk.Button(app, text="Exit", command=app.quit, bg="#dc3545", fg="white").pack(pady=10)

app.mainloop()
